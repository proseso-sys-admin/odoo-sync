"""QAP report builder — merges bill and JE rows, outputs XLSX or DAT.

QAP (Quarterly Alphalist of Payees) groups by partner + ATC code.
One row per partner per ATC per period.
"""

from __future__ import annotations

from typing import BinaryIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from bir_format import qap_dat_line, fmt_date_qap

QAP_COLUMNS = [
    "Taxpayer ID Number",
    "Registered Name",
    "ATC Code",
    "Amount of Income Payment",
    "Amount Tax Withheld",
    "Tax Rate",
    "Source",
]


def build_qap_rows(bill_rows: list[dict], je_rows: list[dict]) -> list[dict]:
    """Merge bill and JE rows, sorted by date ascending."""
    merged = list(bill_rows) + list(je_rows)
    return sorted(merged, key=lambda r: r.get("date", ""))


def write_qap_xlsx(rows: list[dict], output: BinaryIO) -> None:
    """Write QAP rows to an XLSX file in the given output buffer."""
    wb = Workbook()
    ws = wb.active
    ws.title = "QAP"

    header_fill = PatternFill(start_color="293750", end_color="293750", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=10)

    for col_idx, col_name in enumerate(QAP_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row in enumerate(rows, 2):
        values = [
            row.get("tin", ""),
            row.get("registered_name", ""),
            row.get("atc", ""),
            row.get("gross_income", 0),
            row.get("tax_withheld", 0),
            row.get("tax_rate", 0),
            row.get("source", ""),
        ]
        for col_idx, value in enumerate(values, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    wb.save(output)


def write_qap_dat(rows: list[dict]) -> str:
    """Build QAP DAT file content as a string with sequential row numbers."""
    if not rows:
        return ""
    lines = []
    for seq, row in enumerate(rows, 1):
        dat_row = {**row, "date": fmt_date_qap(row["date"])}
        lines.append(qap_dat_line(dat_row, seq=seq))
    return "\r\n".join(lines) + "\r\n"
