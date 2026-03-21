# tests/test_qap_builder.py
import io
import pytest
from openpyxl import load_workbook
from qap_builder import (
    build_qap_rows,
    write_qap_xlsx,
    write_qap_dat,
)


@pytest.fixture
def sample_bill_rows():
    return [
        {
            "tin": "123456789",
            "registered_name": "ACME CORP",
            "last_name": "",
            "first_name": "",
            "middle_name": "",
            "date": "2026-01-15",
            "atc": "WC010",
            "tax_rate": 2,
            "gross_income": 50000.00,
            "tax_withheld": 1000.00,
            "source": "bill",
        },
    ]


@pytest.fixture
def sample_je_rows():
    return [
        {
            "tin": "987654321",
            "registered_name": "VENDOR TWO",
            "last_name": "TWO",
            "first_name": "VENDOR",
            "middle_name": "",
            "date": "2026-02-10",
            "atc": "WI010",
            "tax_rate": 10,
            "gross_income": 30000.00,
            "tax_withheld": 3000.00,
            "source": "journal_entry",
        },
    ]


class TestBuildQapRows:
    def test_merges_and_sorts(self, sample_bill_rows, sample_je_rows):
        merged = build_qap_rows(sample_bill_rows, sample_je_rows)
        assert len(merged) == 2
        assert merged[0]["date"] <= merged[1]["date"]

    def test_empty_jes(self, sample_bill_rows):
        merged = build_qap_rows(sample_bill_rows, [])
        assert len(merged) == 1


class TestWriteQapXlsx:
    def test_produces_valid_xlsx(self, sample_bill_rows, sample_je_rows):
        merged = build_qap_rows(sample_bill_rows, sample_je_rows)
        buf = io.BytesIO()
        write_qap_xlsx(merged, buf)
        buf.seek(0)
        wb = load_workbook(buf)
        ws = wb.active
        assert ws.max_row == 3
        headers = [cell.value for cell in ws[1]]
        assert "ATC Code" in headers
        assert "Amount Tax Withheld" in headers
        assert "Source" in headers


class TestWriteQapDat:
    def test_produces_dat_with_sequence(self, sample_bill_rows, sample_je_rows):
        merged = build_qap_rows(sample_bill_rows, sample_je_rows)
        dat = write_qap_dat(merged)
        lines = dat.strip().split("\r\n")
        assert len(lines) == 2
        assert lines[0].startswith("D1,1601EQ,1,")
        assert lines[1].startswith("D1,1601EQ,2,")
